import json
import os
import glob
import re
from urllib.parse import urlparse, urljoin
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

output_dir = 'katana_outputs'
excel_file = 'katana_multi_scan_report.xlsx'
data = []

target_urls_env = os.environ.get('TARGET_URLS', '')
allowed_domains = set()

for line in target_urls_env.split('\n'):
    line_str = line.strip()
    if line_str:
        domain = re.sub(r'^https?://', '', line_str)
        domain = domain.split('/')[0].split(':')[0]
        allowed_domains.add(domain)

print(f"🎯 [필터 켜짐] 다음 지정 도메인 결과만 엑셀에 저장됩니다: {allowed_domains}")

def calculate_risk_score(url, method):
    score = 0
    url_lower = url.lower()
    
    if method.upper() in ['POST', 'PUT', 'DELETE', 'PATCH']: score += 20
    if '?' in url_lower: score += 30
        
    sensitive_keywords = ['admin', 'login', 'api', 'auth', 'config', 'backup', 'setup', 'manage', 'password', 'token', 'secret', 'dev', 'stg', 'test', 'internal', 'intra', 'debug']
    if any(keyword in url_lower for keyword in sensitive_keywords): score += 40
        
    dangerous_exts = ['.bak', '.env', '.sql', '.zip', '.tar', '.gz', '.log', '.txt', '.xml', '.json', '.swp', '.git']
    if any(url_lower.endswith(ext) or (ext + '?') in url_lower for ext in dangerous_exts): score += 50
        
    script_exts = ['.php', '.jsp', '.asp', '.aspx', '.do', '.action', '.pwkjson']
    if any(url_lower.endswith(ext) or (ext + '?') in url_lower for ext in script_exts): score += 30
        
    static_exts = ['.js', '.css', '.png', '.jpg', '.jpeg', '.gif', '.svg', '.woff', '.woff2', '.ico']
    if any(url_lower.endswith(ext) or (ext + '?') in url_lower for ext in static_exts): score -= 20
        
    final_score = max(score, 0)
    if final_score >= 60: level = 'High'
    elif final_score >= 30: level = 'Medium'
    else: level = 'Low'
        
    return final_score, level

jsonl_files = glob.glob(os.path.join(output_dir, 'part_*.jsonl'))

status_codes_map = {}
if jsonl_files:
    for file_path in jsonl_files:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line_str = line.strip()
                if not line_str: continue
                try:
                    obj = json.loads(line_str)
                    if 'status_code' in obj and 'url' in obj:
                        status_codes_map[obj['url'].rstrip('/')] = str(obj['status_code'])
                except Exception:
                    pass

if jsonl_files:
    for file_path in jsonl_files:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                line_str = line.strip()
                if not line_str: continue
                    
                url_str, method, source, tag, attribute = "", "GET", "", "", ""
                
                try:
                    obj = json.loads(line_str)
                    if 'status_code' in obj: continue
                        
                    if 'request' in obj and isinstance(obj['request'], dict):
                        url_str = obj['request'].get('endpoint') or obj['request'].get('url', '')
                        method = obj['request'].get('method', 'GET')
                        source = obj['request'].get('source') or obj.get('source_url') or obj.get('source', '')
                    else:
                        url_str = obj.get('endpoint') or obj.get('url', '')
                        method = obj.get('method', 'GET')
                        source = obj.get('source_url') or obj.get('source', '')
                        
                    tag = obj.get('tag', '')
                    attribute = obj.get('attribute', '')
                except Exception:
                    pass

                if not url_str:
                    url_match = re.search(r'(https?://[^\s"\'},]+)', line_str)
                    if url_match: url_str = url_match.group(1)

                if not url_str: continue
                
                if not url_str.startswith('http'):
                    if source and source.startswith('http'): url_str = urljoin(source, url_str)
                
                target_source = "Unknown"
                try:
                    parsed_url = urlparse(url_str)
                    if parsed_url.hostname: target_source = parsed_url.hostname
                    else:
                        if '//' in url_str: target_source = url_str.split('/')[2].split(':')[0]
                        else: target_source = url_str.split('/')[0].split(':')[0]
                except Exception: pass
                
                if target_source not in allowed_domains: continue
                
                final_source = "시작 랜딩 페이지(Depth 1)"
                if source and source.strip():
                    clean_src = source.strip()
                    if clean_src.rstrip('/') != url_str.rstrip('/'): final_source = clean_src

                risk_score, risk_level = calculate_risk_score(url_str, method)
                mapped_status = status_codes_map.get(url_str.rstrip('/'), "Error/Timeout")
                    
                row = {
                    '대상 타겟 (Target)': target_source,
                    '응답 상태 (Status)': mapped_status,
                    '위험 등급 (Risk Level)': risk_level,
                    '위험 점수 (Score)': risk_score,
                    '요청 메서드 (Method)': method,
                    '발견된 URL (URL)': url_str,
                    '출처 페이지 (Source)': final_source,
                    'HTML 태그 (Tag)': tag,
                    '속성 (Attribute)': attribute
                }
                data.append(row)

if data:
    df = pd.DataFrame(data)
    df.drop_duplicates(subset=['대상 타겟 (Target)', '발견된 URL (URL)', '요청 메서드 (Method)'], inplace=True)
    df = df.sort_values(by=['위험 점수 (Score)', '대상 타겟 (Target)', '발견된 URL (URL)'], ascending=[False, False, True])
    print(f"🟢 [필터링 완료] 총 {len(df)}개의 고유 타겟 엔드포인트가 엑셀에 매핑되었습니다!")
else:
    df = pd.DataFrame(columns=['대상 타겟 (Target)', '응답 상태 (Status)', '위험 등급 (Risk Level)', '위험 점수 (Score)', '요청 메서드 (Method)', '발견된 URL (URL)', '출처 페이지 (Source)', 'HTML 태그 (Tag)', '속성 (Attribute)'])
    df.loc[0] = ['지정 도메인 내부에서 스캔된 결과가 없거나 차단되었습니다.', '', '', 0, '', '', '', '', '']

wb = Workbook()

font_family = 'Malgun Gothic'
header_font = Font(name=font_family, size=11, bold=True, color='000000')
header_fill = PatternFill(start_color='E6F0FA', end_color='E6F0FA', fill_type='solid')
total_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

high_risk_font = Font(name=font_family, size=10, bold=True, color='FF0000')   
medium_risk_font = Font(name=font_family, size=10, bold=True, color='E26B0A') 
low_risk_font = Font(name=font_family, size=10, color='808080')               
data_font = Font(name=font_family, size=10)

status_200_font = Font(name=font_family, size=10, bold=True, color='0070C0') 
status_400_font = Font(name=font_family, size=10, bold=True, color='E26B0A') 
status_err_font = Font(name=font_family, size=10, color='FF0000')            

center_alignment = Alignment(horizontal='center', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')

ws_dashboard = wb.active
ws_dashboard.title = "대시보드"
ws_dashboard.append(["대상 타겟 도메인 (Target Domain)", "수집된 고유 URL 개수", "상세 페이지 바로가기"])

for cell in ws_dashboard[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
ws_dashboard.row_dimensions[1].height = 24

total_url_count = 0

def apply_styling_and_alignment(ws_sheet, subset_df, start_row=3):
    for row in ws_sheet.iter_rows(min_row=start_row, max_row=ws_sheet.max_row):
        for cell in row:
            if cell.column == 2:
                val = str(cell.value)
                if val.startswith('2'): cell.font = status_200_font
                elif val.startswith('4'): cell.font = status_400_font
                elif val in ['Error/Timeout', 'N/A'] or val.startswith('5'): cell.font = status_err_font
                else: cell.font = data_font
            elif cell.column == 3:
                if cell.value == 'High': cell.font = high_risk_font
                elif cell.value == 'Medium': cell.font = medium_risk_font
                else: cell.font = low_risk_font
            elif cell.column == 4:
                if cell.value >= 60: cell.font = high_risk_font
                elif cell.value >= 30: cell.font = medium_risk_font
                else: cell.font = low_risk_font
            else: cell.font = data_font
                
            if cell.column in [1, 2, 3, 4, 5, 8, 9]: cell.alignment = center_alignment
            else: cell.alignment = left_alignment
            if cell.column in [6, 7]: cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

if not df.empty and len(data) > 0:
    unique_domains = df['대상 타겟 (Target)'].unique()
    last_row_idx = 1
    
    ws_risk = wb.create_sheet(title="🚨 High Risk (우선순위)")
    ws_risk.merge_cells("A1:I1")
    back_btn_risk = ws_risk["A1"]
    back_btn_risk.value = "⬅️ 대시보드 현황판으로 돌아가기"
    back_btn_risk.hyperlink = "#'대시보드'!A1"
    back_btn_risk.font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    back_btn_risk.fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    back_btn_risk.alignment = center_alignment
    ws_risk.row_dimensions[1].height = 26
    
    risk_df = df[df['위험 점수 (Score)'] >= 30].sort_values(by=['위험 점수 (Score)', '대상 타겟 (Target)'], ascending=[False, True])
    if risk_df.empty: risk_df = df.copy()
        
    headers = list(risk_df.columns)
    ws_risk.append(headers)
    for r in dataframe_to_rows(risk_df, index=False, header=False): ws_risk.append(r)
        
    for cell in ws_risk[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    ws_risk.row_dimensions[2].height = 20
    
    # 🔥 [수정 완료] 인자로 risk_df를 정확하게 바인딩하여 데이터 누락 오류를 완벽 패치했습니다.
    apply_styling_and_alignment(ws_risk, risk_df)
                
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(risk_df[col_name].astype(str).map(len).max() if not risk_df.empty else 0, len(str(col_name)))
        if col_idx in [6, 7]: ws_risk.column_dimensions[col_letter].width = min(max_len + 4, 100)
        elif col_idx == 2: ws_risk.column_dimensions[col_letter].width = 16
        else: ws_risk.column_dimensions[col_letter].width = max_len + 3

    for idx, domain in enumerate(unique_domains, start=2):
        sheet_title = domain[:30]
        ws_domain = wb.create_sheet(title=sheet_title)
        
        ws_domain.merge_cells("A1:I1")
        back_btn = ws_domain["A1"]
        back_btn.value = "⬅️ 대시보드 현황판으로 돌아가기"
        back_btn.hyperlink = "#'대시보드'!A1"
        back_btn.font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
        back_btn.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        back_btn.alignment = center_alignment
        ws_domain.row_dimensions[1].height = 26
        
        domain_subset = df[df['대상 타겟 (Target)'] == domain]
        ws_domain.append(headers)
        
        for r in dataframe_to_rows(domain_subset, index=False, header=False): ws_domain.append(r)
            
        for cell in ws_domain[2]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        ws_domain.row_dimensions[2].height = 20
            
        apply_styling_and_alignment(ws_domain, domain_subset)
                    
        for col_idx, col_name in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(domain_subset[col_name].astype(str).map(len).max() if not domain_subset.empty else 0, len(str(col_name)))
            if col_idx in [6, 7]: ws_domain.column_dimensions[col_letter].width = min(max_len + 4, 100)
            elif col_idx == 2: ws_domain.column_dimensions[col_letter].width = 16
            else: ws_domain.column_dimensions[col_letter].width = max_len + 3
            
        current_domain_count = len(domain_subset)
        total_url_count += current_domain_count
        
        ws_dashboard.cell(row=idx, column=1, value=domain).alignment = center_alignment
        ws_dashboard.cell(row=idx, column=2, value=current_domain_count).alignment = center_alignment
        
        link_cell = ws_dashboard.cell(row=idx, column=3, value="🔍 상세 내역 시트로 탭 이동")
        link_cell.hyperlink = f"#'{sheet_title}'!A2"
        link_cell.font = Font(name=font_family, size=10, color='0056B3', underline='single')
        link_cell.alignment = center_alignment
        
        ws_dashboard.row_dimensions[idx].height = 22
        ws_dashboard.cell(row=idx, column=1).font = data_font
        ws_dashboard.cell(row=idx, column=2).font = Font(name=font_family, size=10, bold=True)
        last_row_idx = idx

    ws_dashboard.insert_rows(2)
    ws_dashboard.cell(row=2, column=1, value="🚨 [전체 통합] 취약점 점검 최우선순위 (High Risk)").alignment = center_alignment
    ws_dashboard.cell(row=2, column=1).font = Font(name=font_family, size=10, bold=True, color='C00000')
    ws_dashboard.cell(row=2, column=2, value=len(df[df['위험 점수 (Score)'] >= 30])).alignment = center_alignment
    ws_dashboard.cell(row=2, column=2).font = Font(name=font_family, size=11, bold=True, color='C00000')
    link_risk = ws_dashboard.cell(row=2, column=3, value="🔥 High Risk 시트로 이동")
    link_risk.hyperlink = f"#'🚨 High Risk (우선순위)'!A2"
    link_risk.font = Font(name=font_family, size=10, bold=True, color='C00000', underline='single')
    link_risk.alignment = center_alignment
    ws_dashboard.row_dimensions[2].height = 24
    
    total_row_idx = last_row_idx + 2
    ws_dashboard.cell(row=total_row_idx, column=1, value="📊 수집된 URL 총합 (Total)").alignment = center_alignment
    ws_dashboard.cell(row=total_row_idx, column=1).font = Font(name=font_family, size=10, bold=True, color='FF0000')
    ws_dashboard.cell(row=total_row_idx, column=1).fill = total_fill
    
    ws_dashboard.cell(row=total_row_idx, column=2, value=total_url_count).alignment = center_alignment
    ws_dashboard.cell(row=total_row_idx, column=2).font = Font(name=font_family, size=11, bold=True, color='FF0000')
    ws_dashboard.cell(row=total_row_idx, column=2).fill = total_fill
    
    ws_dashboard.cell(row=total_row_idx, column=3, value=\"\").fill = total_fill
    ws_dashboard.row_dimensions[total_row_idx].height = 24

else:
    ws_empty = wb.create_sheet(title="결과 없음")
    ws_empty.append(list(df.columns))
    ws_empty.append(['지정 도메인 내부에서 스캔된 결과가 없거나 차단되었습니다.', '', '', 0, '', '', '', '', ''])
    
    ws_dashboard.cell(row=2, column=1, value="N/A").alignment = center_alignment
    ws_dashboard.cell(row=2, column=2, value=0).alignment = center_alignment
    ws_dashboard.cell(row=2, column=3, value="데이터 없음").alignment = center_alignment

ws_dashboard.column_dimensions['A'].width = 45
ws_dashboard.column_dimensions['B'].width = 24
ws_dashboard.column_dimensions['C'].width = 30

wb.save(excel_file)
print(f"🏁 [최적화 완료] 엑셀 데이터 매핑 리포트 출력 세이브 완료: {excel_file}")
